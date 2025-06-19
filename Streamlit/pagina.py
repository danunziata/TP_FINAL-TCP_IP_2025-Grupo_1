import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from datetime import datetime, timedelta, time
import os
from influxdb_client import InfluxDBClient
import json
import subprocess
import yaml
from yaml.loader import SafeLoader
from pathlib import Path
import bcrypt
import time as time_module
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

# --- Configuración de InfluxDB desde variables de entorno ---
INFLUX_URL = os.getenv('INFLUXDB_URL', 'http://influxdb:8086')
INFLUX_TOKEN = os.getenv('INFLUXDB_TOKEN', 'token_telegraf')
INFLUX_ORG = os.getenv('INFLUXDB_ORG', 'power_logic')
INFLUX_BUCKET = os.getenv('INFLUXDB_BUCKET', 'mensualx6')

# --- Clientes de archivos de configuración y datos ---
CONFIG_FILE_PATH = Path(__file__).parent / "config.yaml"
USUARIOS_FILE_PATH = Path(__file__).parent / "usuarios.json"
RESET_TOKENS_FILE_PATH = Path(__file__).parent / "reset_tokens.json"

# --- Funciones auxiliares para cargar/guardar archivos ---
def load_config():
    """Carga el archivo config.yaml."""
    try:
        if not CONFIG_FILE_PATH.exists():
            st.error(f"Error: El archivo de configuración '{CONFIG_FILE_PATH}' no existe.")
            return None
        with CONFIG_FILE_PATH.open("r", encoding="utf-8") as f:
            return yaml.load(f, Loader=SafeLoader)
    except Exception as e:
        st.error(f"Error al cargar config.yaml: {e}")
        return None

def save_config(config_data):
    """Guarda el archivo config.yaml."""
    try:
        with CONFIG_FILE_PATH.open("w", encoding="utf-8") as f:
            yaml.dump(config_data, f, default_flow_style=False)
        return True
    except Exception as e:
        st.error(f"Error al guardar config.yaml: {e}")
        return False

def load_usuarios():
    """Carga el archivo usuarios.json."""
    if os.path.exists(USUARIOS_FILE_PATH):
        try:
            with USUARIOS_FILE_PATH.open("r", encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:
            st.error(f"Error al cargar usuarios.json: {e}")
            return []
    return []

def save_usuarios(usuarios_data):
    """Guarda el archivo usuarios.json."""
    try:
        with USUARIOS_FILE_PATH.open("w", encoding="utf-8") as f:
            json.dump(usuarios_data, f, indent=4)
        return True
    except Exception as e:
        st.error(f"Error al guardar usuarios.json: {e}")
        return False

# ... (El resto de las funciones de carga/guardado de JSON permanecen igual) ...

# --- Cliente de InfluxDB y Carga de Datos ---
def get_influx_client():
    return InfluxDBClient(url=INFLUX_URL, token=INFLUX_TOKEN, org=INFLUX_ORG)

@st.cache_data
def load_data(_file_mod=None, data_version=0):
    """Carga datos desde InfluxDB, aplica escala y el mapeo de nombres solicitado."""
    client = get_influx_client()
    query_api = client.query_api()
    flux = f'''
    from(bucket: "{INFLUX_BUCKET}")
    |> range(start: -30d)
    |> filter(fn: (r) => r["_measurement"] == "modbus")
    |> filter(fn: (r) => r["host"] == "telegraf")
    |> pivot(rowKey:["_time"], columnKey:["_field"], valueColumn:"_value")
    '''
    try:
        result = query_api.query_data_frame(flux)
        if result is None or result.empty:
            return pd.DataFrame()

        df = result.copy()
        if isinstance(df, list): df = pd.concat(df, ignore_index=True)

        df.drop(columns=['result', 'table', '_start', '_stop', '_measurement', 'host'], errors='ignore', inplace=True)
        df.rename(columns={'_time': 'timestamp'}, inplace=True)
        df['timestamp'] = pd.to_datetime(df['timestamp']).dt.tz_localize(None)

        # Aplicar escala a los nombres TÉCNICOS (antes de renombrar)
        # Basado en telegraf.conf, estas columnas tienen una escala de 0.1
        cols_to_scale = ['voltaje_a_n', 'voltaje_b_n', 'voltaje_c_n', 'corriente_a', 'corriente_b', 'corriente_c']
        for col in cols_to_scale:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce') / 10.0

        # Aplicar el mapeo de nombres solicitado por el usuario
        field_mapping = {
            'corriente_a': 'Corriente A (A)', 'corriente_b': 'Corriente B (A)', 'corriente_c': 'Corriente C (A)',
            'voltaje_a_n': 'Voltaje A-N (V)', 'voltaje_b_n': 'Voltaje B-N (V)', 'voltaje_c_n': 'Voltaje C-N (V)',
            'potencia_activa_a': 'Potencia Activa A (W)', 'potencia_activa_b': 'Potencia Activa B (W)', 'potencia_activa_c': 'Potencia Activa C (W)',
            'potencia_activa_total': 'Potencia Activa Total (W)', 'potencia_reactiva_a': 'Potencia Reactiva A (VAR)',
            'potencia_reactiva_b': 'Potencia Reactiva B (VAR)', 'potencia_reactiva_c': 'Potencia Reactiva C (VAR)',
            'potencia_reactiva_total': 'Potencia Reactiva Total (VAR)', 'potencia_aparente_a': 'Potencia Aparente A (VA)',
            'potencia_aparente_b': 'Potencia Aparente B (VA)', 'potencia_aparente_c': 'Potencia Aparente C (VA)',
            'potencia_aparente_total': 'Potencia Aparente Total (VA)',
            'demanda_potencia_real_3_fases_running': 'Demanda Potencia Real (W)'
        }
        df.rename(columns=field_mapping, inplace=True)
        
        return df

    except Exception as e:
        st.error(f"Error en load_data: {e}")
        return pd.DataFrame()

# --- Funciones Mejoradas del Dashboard ---
def detect_columns(df):
    """Detecta y categoriza columnas, ignorando 'timestamp'."""
    if df is None or df.empty:
        return {}, {}
    
    columns_to_scan = [col for col in df.columns if col.lower() != 'timestamp']
    
    # CORRECCIÓN: Se añaden patrones en inglés y español para ser más robustos.
    detected = {
        'Voltajes': [col for col in columns_to_scan if any(p in col.lower() for p in ['voltaje', 'voltage'])],
        'Corrientes': [col for col in columns_to_scan if any(p in col.lower() for p in ['corriente', 'current'])],
        'Potencia Activa': [col for col in columns_to_scan if any(p in col.lower() for p in ['potencia activa', 'active_power'])],
        'Potencia Reactiva': [col for col in columns_to_scan if any(p in col.lower() for p in ['potencia reactiva', 'reactive_power'])],
        'Potencia Aparente': [col for col in columns_to_scan if any(p in col.lower() for p in ['potencia aparente', 'apparent_power'])],
        'Demanda': [col for col in columns_to_scan if any(p in col.lower() for p in ['demanda', 'demand'])]
    }
    # Filtrar categorías que no encontraron ninguna columna
    detected = {k: v for k, v in detected.items() if v}
    
    friendly_names = {col: col for col in df.columns if col != 'timestamp'}
    return detected, friendly_names

def create_multi_series_chart(data, title, y_columns, y_title, colors=None):
    """Crea gráficos. Los nombres de las series ya serán legibles."""
    fig = go.Figure()
    if colors is None:
        colors = ['#FF4B4B', '#0068C9', '#00C39F', '#FF8C00', '#9467BD']
    
    series_added = 0
    for i, col in enumerate(y_columns):
        if col in data.columns and not data[col].dropna().empty:
            fig.add_trace(go.Scatter(x=data['timestamp'], y=data[col], mode='lines', name=col,
                                     line=dict(width=2, color=colors[i % len(colors)])))
            series_added += 1
    
    if series_added > 0:
        fig.update_layout(title=title, xaxis_title='Fecha y Hora', yaxis_title=y_title,
                          template='plotly_white', hovermode='x unified',
                          legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='right', x=1))
    return fig, series_added

def create_metrics_dashboard(data, detected_columns):
    st.subheader("📈 Resumen de Métricas")
    cols = st.columns(4)
    metric_idx = 0
    def display_metric(label, value, unit):
        nonlocal metric_idx
        if metric_idx < 4:
            cols[metric_idx].metric(label, f"{value:.2f} {unit}")
            metric_idx += 1

    if 'Voltajes' in detected_columns:
        volt_data = data[detected_columns['Voltajes']].select_dtypes(include=np.number)
        if not volt_data.empty: display_metric("Voltaje Promedio", volt_data.mean().mean(), "V")
    
    if 'Corrientes' in detected_columns:
        curr_data = data[detected_columns['Corrientes']].select_dtypes(include=np.number)
        if not curr_data.empty: display_metric("Corriente Promedio", curr_data.mean().mean(), "A")

    if 'Potencia Activa' in detected_columns:
        power_data = data[detected_columns['Potencia Activa']].select_dtypes(include=np.number)
        if not power_data.empty: display_metric("Pot. Activa Prom", power_data.mean().mean(), "W")
    
    display_metric("Registros", len(data), "")


def create_excel_file(data, detected_columns):
    """Crea un archivo Excel avanzado. Las columnas ya tienen nombres legibles."""
    buffer = BytesIO()
    wb = Workbook()
    
    # Hoja 1: Datos Completos
    ws_data = wb.active
    ws_data.title = "Datos Completos"
    for r in dataframe_to_rows(data, index=False, header=True):
        ws_data.append(r)

    # Hoja 2: Resumen Estadístico
    ws_summary = wb.create_sheet(title="Resumen Estadistico")
    summary_data = []
    numeric_cols = data.select_dtypes(include=np.number).columns
    for col in numeric_cols:
        col_data = data[col].dropna()
        if not col_data.empty:
            stats = {'Metrica': col, 'Promedio': col_data.mean(), 'Maximo': col_data.max(),
                     'Minimo': col_data.min(), 'Desv. Estandar': col_data.std(),
                     'Registros': col_data.count()}
            summary_data.append(stats)
    if summary_data:
        df_summary = pd.DataFrame(summary_data)
        for r in dataframe_to_rows(df_summary, index=False, header=True):
            ws_summary.append(r)

    # Hojas por Categoría
    for category, cols in detected_columns.items():
        if cols:
            df_category = data[['timestamp'] + cols]
            sheet_name = category.replace('_', ' ').title()[:31]
            ws_cat = wb.create_sheet(title=sheet_name)
            for r in dataframe_to_rows(df_category, index=False, header=True):
                ws_cat.append(r)

    # Autoajuste de columnas
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column].width = adjusted_width

    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# --- EL RESTO DEL CÓDIGO (ALERTAS, USUARIOS, MAIN) PERMANECE IGUAL ---
# ... (Aquí irían todas las demás funciones como mostrar_alertas_activas, display_franja_config_form, etc., que no necesitan cambios)
# PEGAR AQUÍ EL RESTO DE LAS FUNCIONES DE LA VERSIÓN ANTERIOR
def load_reset_tokens_file():
    """Carga los tokens de restablecimiento desde el archivo JSON."""
    if os.path.exists(RESET_TOKENS_FILE_PATH):
        try:
            with RESET_TOKENS_FILE_PATH.open("r", encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:
            st.warning(f"Error al cargar reset_tokens.json: {e}. Se iniciará con tokens vacíos.")
            return {}
    return {}

def save_reset_tokens_file(tokens_data):
    """Guarda los tokens de restablecimiento en el archivo JSON."""
    try:
        with RESET_TOKENS_FILE_PATH.open("w", encoding="utf-8") as f:
            json.dump(tokens_data, f, indent=4)
        return True
    except Exception as e:
        st.error(f"Error al guardar reset_tokens.json: {e}")
        return False
        
def mostrar_alertas_activas(logs_file="logs_alertas.json"):
    st.subheader("📋 Historial de Alertas")
    try:
        if not os.path.exists(logs_file):
            st.info("Aún no se han registrado alertas.")
            return pd.DataFrame()

        with open(logs_file, "r", encoding="utf-8") as f:
            logs = json.load(f)
        
        if not logs:
            st.info("El historial de alertas está vacío.")
            return pd.DataFrame()

        df_logs = pd.DataFrame(logs)
        df_logs['timestamp'] = pd.to_datetime(df_logs['timestamp'], errors='coerce')
        df_logs.dropna(subset=['timestamp'], inplace=True)
        df_logs.sort_values(by='timestamp', ascending=False, inplace=True)

        st.sidebar.header("Filtrar Alertas")
        fecha_minima = df_logs['timestamp'].min().date()
        fecha_maxima = df_logs['timestamp'].max().date()

        fecha_inicio = st.sidebar.date_input("Fecha de Inicio", value=fecha_minima, min_value=fecha_minima, max_value=fecha_maxima, key="filtro_fecha_inicio")
        fecha_fin = st.sidebar.date_input("Fecha de Fin", value=fecha_maxima, min_value=fecha_minima, max_value=fecha_maxima, key="filtro_fecha_fin")

        if fecha_inicio > fecha_fin:
            st.sidebar.error("Error: La fecha de inicio no puede ser posterior a la fecha de fin.")
            return pd.DataFrame()

        fecha_inicio_dt = datetime.combine(fecha_inicio, time.min)
        fecha_fin_dt = datetime.combine(fecha_fin, time.max)
        
        df_filtrado = df_logs[(df_logs['timestamp'] >= fecha_inicio_dt) & (df_logs['timestamp'] <= fecha_fin_dt)]
        
        st.write(f"Mostrando {len(df_filtrado)} de {len(df_logs)} alertas totales.")
        df_filtrado_vista = df_filtrado.reset_index(drop=True)
        st.dataframe(df_filtrado_vista, use_container_width=True)

        if not df_filtrado.empty:
            csv = df_filtrado.to_csv(index=False).encode('utf-8')
            st.download_button(
                label="📥 Descargar historial filtrado (CSV)",
                data=csv,
                file_name=f"historial_alertas_{fecha_inicio}_a_{fecha_fin}.csv",
                mime="text/csv",
            )
        return df_filtrado

    except Exception as e:
        st.error(f"Error al leer el historial de alertas: {e}")
        return pd.DataFrame()

display_names = {'voltaje': 'Voltaje', 'current_l1': 'Corriente L1', 'active_power': 'Potencia Activa'}
default_thresholds_for_display = {'voltaje': {'min': 0.0, 'max': 0.0}, 'current_l1': {'min': 0.0, 'max': 0.0}, 'active_power': {'min': 0.0, 'max': 0.0}}

def display_franja_config_form(current_config, franjas_horarias_data):
    st.subheader("⚙️ Configuración de Umbrales por Franja Horaria")
    
    if not franjas_horarias_data:
        st.warning("No se encontraron franjas horarias configuradas.")
        return
    
    franjas_names = list(franjas_horarias_data.keys())
    
    if not franjas_names:
        st.error("No hay franjas horarias definidas.")
        return

    selected_franja = st.selectbox("Seleccionar Franja Horaria:", options=franjas_names, key="franja_horaria_selector")
    current_franja_details = franjas_horarias_data.get(selected_franja, {})
    current_umbrales = current_franja_details.get("umbrales", {})
    es_admin = st.session_state.get('roles') == 'admin'

    st.markdown(f"**Horario de la franja '{selected_franja}'**")
    col_inicio_hora, col_fin_hora = st.columns(2)
    
    try: initial_inicio_time = datetime.strptime(current_franja_details.get('inicio_hora', '00:00'), "%H:%M").time()
    except ValueError: initial_inicio_time = time(0, 0)
    
    try: initial_fin_time = datetime.strptime(current_franja_details.get('fin_hora', '00:00'), "%H:%M").time()
    except ValueError: initial_fin_time = time(0, 0)

    with col_inicio_hora:
        new_inicio_time = st.time_input("Hora de Inicio:", value=initial_inicio_time, key=f"{selected_franja}_inicio_time_input", disabled=not es_admin)
    with col_fin_hora:
        new_fin_time = st.time_input("Hora de Fin:", value=initial_fin_time, key=f"{selected_franja}_fin_time_input", disabled=not es_admin)

    st.markdown(f"Ajusta los umbrales para **{selected_franja}**.")
    nuevos_umbrales_para_franja = {}
    for variable_key, defaults in default_thresholds_for_display.items():
        display_name = display_names.get(variable_key, variable_key)
        st.write(f"**{display_name}**")
        col_min, col_max = st.columns(2)
        initial_min = current_umbrales.get(variable_key, {}).get('min', defaults['min'])
        initial_max = current_umbrales.get(variable_key, {}).get('max', defaults['max'])
        with col_min:
            new_min = st.number_input(f"Mínimo para {display_name}", value=float(initial_min), format="%.2f", key=f"{selected_franja}_{variable_key}_min_input", disabled=not es_admin)
        with col_max:
            new_max = st.number_input(f"Máximo para {display_name}", value=float(initial_max), format="%.2f", key=f"{selected_franja}_{variable_key}_max_input", disabled=not es_admin)
        nuevos_umbrales_para_franja[variable_key] = {'min': new_min, 'max': new_max}

    st.divider()
    
    if es_admin:
        st.subheader("✉️ Configuración de Notificaciones Generales")
        notificaciones_activas = current_config.get("notificaciones_generales", False)
        current_digest_interval = current_config.get("alert_digest_interval_minutes", 1440)
        
        activar_mail_general = st.checkbox("**Activar envío de correos de alerta (global)**", value=notificaciones_activas)
        new_digest_interval = st.number_input("Frecuencia de Envío de Resumen (minutos):", min_value=1, value=current_digest_interval, key="alert_digest_interval_input")
        
        if st.button("💾 Guardar Configuración", use_container_width=True):
            franjas_a_guardar = current_config.get('franjas_horarias', {}).copy()
            franjas_a_guardar[selected_franja] = {
                'inicio_hora': new_inicio_time.strftime("%H:%M"),
                'fin_hora': new_fin_time.strftime("%H:%M"),
                'umbrales': nuevos_umbrales_para_franja
            }
            current_config['franjas_horarias'] = franjas_a_guardar
            current_config['notificaciones_generales'] = activar_mail_general
            current_config['alert_digest_interval_minutes'] = new_digest_interval
            
            if save_config(current_config):
                st.success("✅ Configuración actualizada.")
                time_module.sleep(2)
                st.rerun()
            else:
                st.error("❌ Error al guardar la configuración.")
    else:
        st.info("ℹ️ Como usuario normal, solo puedes visualizar esta configuración.")

def wizard_configuracion_franjas(current_config):
    st.title("Asistente de Configuración Inicial de Franjas Horarias")
    st.warning("Parece que es la primera vez que configuras las franjas horarias.")
    default_franjas = current_config.get("default_franjas_horarias", {})
    if not default_franjas:
        st.error("Error: 'default_franjas_horarias' no encontrado en 'config.yaml'.")
        return

    st.subheader("Definición de Franjas")
    col1, col2, col3, col4 = st.columns(4)
    with col1: dia_inicio = st.time_input("Inicio DIA", value=datetime.strptime(default_franjas['DIA']['inicio_hora'], "%H:%M").time(), key="wiz_dia_inicio")
    with col2: dia_fin = st.time_input("Fin DIA", value=datetime.strptime(default_franjas['DIA']['fin_hora'], "%H:%M").time(), key="wiz_dia_fin")
    with col3: noche_inicio = st.time_input("Inicio NOCHE", value=datetime.strptime(default_franjas['NOCHE']['inicio_hora'], "%H:%M").time(), key="wiz_noche_inicio")
    with col4: noche_fin = st.time_input("Fin NOCHE", value=datetime.strptime(default_franjas['NOCHE']['fin_hora'], "%H:%M").time(), key="wiz_noche_fin")

    nuevas_franjas_config = {}
    st.subheader("Umbrales por Defecto")
    for franja_name, franja_defaults in default_franjas.items():
        st.markdown(f"**Umbrales para {franja_name}**")
        umbrales = {}
        for var_key, default_vals in default_thresholds_for_display.items():
            display_name = display_names.get(var_key, var_key)
            c1, c2 = st.columns(2)
            min_val = c1.number_input(f"Mínimo {display_name} ({franja_name})", value=float(franja_defaults['umbrales'].get(var_key, {}).get('min', 0.0)), format="%.2f", key=f"wiz_{franja_name}_{var_key}_min")
            max_val = c2.number_input(f"Máximo {display_name} ({franja_name})", value=float(franja_defaults['umbrales'].get(var_key, {}).get('max', 0.0)), format="%.2f", key=f"wiz_{franja_name}_{var_key}_max")
            umbrales[var_key] = {'min': min_val, 'max': max_val}
        
        if franja_name == 'DIA':
            nuevas_franjas_config['DIA'] = {'inicio_hora': dia_inicio.strftime("%H:%M"), 'fin_hora': dia_fin.strftime("%H:%M"), 'umbrales': umbrales}
        else:
            nuevas_franjas_config['NOCHE'] = {'inicio_hora': noche_inicio.strftime("%H:%M"), 'fin_hora': noche_fin.strftime("%H:%M"), 'umbrales': umbrales}

    if st.button("💾 Guardar Configuración Inicial", use_container_width=True, type="primary"):
        current_config['franjas_horarias'] = nuevas_franjas_config
        if save_config(current_config):
            st.success("✅ Franjas configuradas. Puedes continuar.")
            time_module.sleep(2)
            st.rerun()
        else:
            st.error("❌ Error al guardar la configuración.")

def ejecutar_checker_manual():
    st.subheader("🔍 Ejecutar Análisis Manual de Alertas")
    if st.button("🚨 Ejecutar Checker Ahora", use_container_width=True, type="primary"):
        with st.spinner('Ejecutando checker.py...'):
            try:
                result = subprocess.run(["python", "checker.py", "--manual-run"], capture_output=True, text=True, check=True)
                st.success("¡Análisis completado!")
                st.code(result.stdout)
                if result.stderr: st.warning("Advertencias:\n" + result.stderr)
            except Exception as e:
                st.error(f"Error al ejecutar checker.py: {e}")

def editar_perfil_usuario():
    st.title("👤 Editar Perfil de Usuario")
    usuario_login_email = st.session_state.get("email")
    if not usuario_login_email:
        st.warning("No se pudo detectar el email del usuario.")
        return

    usuarios = load_usuarios()
    usuario_actual_data = next((u for u in usuarios if u.get("login_email") == usuario_login_email), None)
    if not usuario_actual_data:
        st.warning(f"Tu email ({usuario_login_email}) no está en 'usuarios.json'.")
        return

    st.info(f"Editando perfil para: **{st.session_state.get('username')}**")
    with st.form("edit_profile_form"):
        nuevo_nombre = st.text_input("Nombre:", value=usuario_actual_data.get("nombre", ""))
        nuevo_email_alertas = st.text_input("Email para alertas:", value=usuario_actual_data.get("alert_email", usuario_login_email))
        nueva_pref_notificacion = st.checkbox("✅ Recibir alertas por email", value=usuario_actual_data.get("recibir_notificaciones", False))
        
        if st.form_submit_button("💾 Guardar Cambios"):
            idx = usuarios.index(usuario_actual_data)
            usuarios[idx]["nombre"] = nuevo_nombre
            usuarios[idx]["alert_email"] = nuevo_email_alertas
            usuarios[idx]["recibir_notificaciones"] = nueva_pref_notificacion
            if save_usuarios(usuarios):
                st.success("Perfil actualizado.")
                time_module.sleep(2)
                st.rerun()
            else:
                st.error("❌ Error al guardar perfil.")

    st.subheader("Cambiar Contraseña")
    with st.form("change_password_form"):
        current_password = st.text_input("Contraseña Actual", type="password")
        new_password = st.text_input("Nueva Contraseña", type="password")
        confirm_new_password = st.text_input("Confirmar Nueva Contraseña", type="password")
        if st.form_submit_button("Cambiar Contraseña"):
            if new_password != confirm_new_password:
                st.error("Las contraseñas nuevas no coinciden.")
                return
            
            current_config = load_config()
            username = st.session_state.get('username')
            user_creds = current_config['credentials']['usernames'].get(username)
            
            if user_creds and bcrypt.checkpw(current_password.encode('utf-8'), user_creds['password'].encode('utf-8')):
                new_hashed = bcrypt.hashpw(new_password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
                current_config['credentials']['usernames'][username]['password'] = new_hashed
                if save_config(current_config):
                    st.success("Contraseña cambiada. Serás redirigido al login.")
                    st.session_state['authentication_status'] = None # Forzar logout
                    time_module.sleep(2)
                    st.rerun()
                else:
                    st.error("Error al guardar la nueva contraseña.")
            else:
                st.error("La contraseña actual es incorrecta.")

def gestionar_usuarios():
    st.title("👥 Gestión de Usuarios")
    if st.session_state.get('roles') != 'admin':
        st.warning("🚫 Acceso denegado.")
        return

    config_data = load_config()
    usernames_in_config = config_data['credentials']['usernames']
    df_users = pd.DataFrame.from_dict(usernames_in_config, orient='index').reset_index()
    df_users = df_users.rename(columns={'index': 'Username'})
    st.dataframe(df_users[['Username', 'email', 'first_name', 'last_name', 'roles']], use_container_width=True)

    st.subheader("Eliminar Usuario")
    user_to_delete = st.selectbox("Selecciona un usuario a eliminar:", options=[""] + [u for u in usernames_in_config.keys() if u != st.session_state.get('username')])
    if user_to_delete and st.button(f"🔴 Eliminar {user_to_delete}", type="secondary"):
        deleted_email = config_data['credentials']['usernames'].pop(user_to_delete, {}).get('email')
        if save_config(config_data):
            usuarios_data = load_usuarios()
            usuarios_data = [u for u in usuarios_data if u.get('login_email') != deleted_email]
            save_usuarios(usuarios_data)
            st.success(f"Usuario '{user_to_delete}' eliminado.")
            time_module.sleep(2)
            st.rerun()
        else:
            st.error("Error al eliminar el usuario.")

def crear_dashboard_alertas(df_alertas):
    st.divider()
    st.subheader("📊 Dashboard de Alertas")
    if df_alertas.empty:
        st.info("No hay alertas para mostrar gráficos.")
        return

    col1, col2 = st.columns(2)
    with col1:
        # La columna es 'conteo' a partir de pandas 2.x
        dist_variable = df_alertas['variable'].value_counts().reset_index()
        fig_pie_var = go.Figure(go.Pie(labels=dist_variable['variable'], values=dist_variable['count'], hole=.3, title='Alertas por Variable'))
        st.plotly_chart(fig_pie_var, use_container_width=True)
    with col2:
        if 'franja_horaria' in df_alertas.columns:
            dist_franja = df_alertas['franja_horaria'].value_counts().reset_index()
            fig_pie_franja = go.Figure(go.Pie(labels=dist_franja['franja_horaria'], values=dist_franja['count'], hole=.3, title='Alertas por Franja'))
            st.plotly_chart(fig_pie_franja, use_container_width=True)


def main():
    if 'data_version' not in st.session_state: st.session_state.data_version = 0
    if 'last_update' not in st.session_state: st.session_state.last_update = datetime.now()
    if st.session_state.get('email') is None:
        st.error("No hay sesión de usuario. Por favor, inicia sesión.")
        st.stop()

    current_config = load_config()
    if not current_config:
        st.error("Error crítico: No se pudo cargar config.yaml.")
        st.stop()

    if st.session_state.get('roles') == 'admin' and not current_config.get('franjas_horarias'):
        wizard_configuracion_franjas(current_config)
        st.stop()

    st.sidebar.title("Menú Principal")
    nav_options = ["📈 Dashboard Principal", "🔔 Alertas y Logs", "👤 Editar Perfil"]
    if st.session_state.get('roles') == 'admin':
        nav_options.append("👥 Gestión de Usuarios")
    seccion = st.sidebar.radio("Navegar:", nav_options, index=0)

    # --- INICIO DE LA SECCIÓN DEL DASHBOARD (LÓGICA MEJORADA) ---
    if seccion == "📈 Dashboard Principal":
        st.title("📊 Sistema de Monitoreo PowerLogic 4000")
        st.markdown("**Visualización completa de parámetros eléctricos** | Schneider Electric™")
        st.divider()

        with st.spinner('Cargando y procesando datos...'):
            df = load_data(data_version=st.session_state.data_version)

        if df is None or df.empty:
            st.error("❌ No se pudieron cargar los datos de InfluxDB.")
            st.info("Verifica la conexión y que el bucket contenga datos.")
            if st.button("🔄 Reintentar conexión"): st.rerun()
            st.stop()

        detected_columns, friendly_names = detect_columns(df)
        st.session_state.last_update = datetime.now()

        with st.sidebar:
            st.header("⚙️ Configuración del Dashboard")
            if st.button("🔄 Actualizar Datos", use_container_width=True):
                with st.spinner('Actualizando...'):
                    st.session_state.data_version += 1
                    load_data.clear()
                    st.rerun()
            st.caption(f"🕒 Última actualización: {st.session_state.last_update.strftime('%H:%M:%S')}")
            st.divider()
            
            min_d, max_d = df['timestamp'].min().date(), df['timestamp'].max().date()
            date_range = st.date_input("Seleccionar rango de fechas:", value=(min_d, max_d), min_value=min_d, max_value=max_d)
            
            st.subheader("Métricas a visualizar:")
            show_sections = {}
            for category in detected_columns.keys():
                show_sections[category] = st.checkbox(category, value=True)
            
        if len(date_range) == 2:
            start, end = date_range
            mask = (df['timestamp'] >= pd.Timestamp(start)) & (df['timestamp'] < pd.Timestamp(end) + timedelta(days=1))
            filtered_df = df[mask]
            
            if filtered_df.empty:
                st.warning("⚠️ No hay datos para el rango seleccionado.")
            else:
                create_metrics_dashboard(filtered_df, detected_columns)
                st.divider()

                for category, cols in detected_columns.items():
                    if show_sections.get(category):
                        st.subheader(f"⚡ {category}")
                        unit = "V" if "Voltaje" in category else "A" if "Corriente" in category else "W" if "Potencia" in category or "Demanda" in category else "VA" if "Aparente" in category else "VAR" if "Reactiva" in category else ""
                        chart, count = create_multi_series_chart(filtered_df, category, cols, f"{unit}")
                        if count > 0: st.plotly_chart(chart, use_container_width=True)
                        else: st.info(f"No hay datos de {category.lower()} válidos para mostrar.")
                
                st.divider()
                st.subheader("📊 Datos Detallados")
                st.dataframe(filtered_df, height=400, use_container_width=True)
                
                st.subheader("📥 Exportar Datos")
                col1_exp, col2_exp = st.columns(2)
                with col1_exp:
                    csv_data = filtered_df.to_csv(index=False).encode('utf-8')
                    st.download_button("📄 Descargar CSV", data=csv_data, file_name=f"powerlogic_data_{start}_{end}.csv", mime="text/csv", use_container_width=True)
                
                with col2_exp:
                    excel_data = create_excel_file(filtered_df, detected_columns)
                    st.download_button("📊 Descargar Excel Avanzado", data=excel_data, file_name=f"powerlogic_completo_{start}_{end}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
        else:
            st.error("❌ Por favor selecciona un rango de fechas válido")

    elif seccion == "🔔 Alertas y Logs":
        st.title("🛎️ Sistema de Alertas y Logs")
        st.markdown("**Administra umbrales y revisa el historial de alertas**")
        st.divider()
        df_alertas_filtradas = mostrar_alertas_activas()
        if df_alertas_filtradas is not None and not df_alertas_filtradas.empty:
            crear_dashboard_alertas(df_alertas_filtradas)
        st.divider()
        config_data = load_config()
        if config_data:
            franjas = config_data.get('franjas_horarias') or config_data.get('default_franjas_horarias', {})
            display_franja_config_form(config_data, franjas)
        st.divider()
        ejecutar_checker_manual()

    elif seccion == "👤 Editar Perfil":
        editar_perfil_usuario()
    
    elif seccion == "👥 Gestión de Usuarios":
        gestionar_usuarios()

    st.divider()
    st.caption(f"© {datetime.today().year} Schneider Electric - Power Monitoring System | v3.0 | Mapeo Integrado")

if __name__ == "__main__":
    main()
